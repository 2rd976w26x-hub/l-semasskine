"""Excel -> JSON parser for Læsemaskine.

Usage:
  python excel_to_json.py --excel ../shared/Ordtraening.xlsx --out ../data/words.json

Note:
- Keeps a slim word object for gameplay plus a 'raw' dict with all columns.
"""

from __future__ import annotations
import argparse
import json
import datetime
from pathlib import Path
import re
import openpyxl

def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--excel", required=True, help="Path to Excel file (.xlsx)")
    p.add_argument("--out", required=True, help="Output JSON path")
    p.add_argument("--sheet", default=None, help="Sheet name (default: first sheet)")
    p.add_argument("--version", default="0.1.0", help="Data version stamp")
    return p.parse_args()

def main():
    args = parse_args()
    excel_path = Path(args.excel).resolve()
    out_path = Path(args.out).resolve()

    wb = openpyxl.load_workbook(excel_path)
    sheet = wb[args.sheet] if args.sheet else wb.active

    headers = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows = []
    for r in sheet.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        d = {}
        for i, h in enumerate(headers):
            key = h if h is not None else f"col{i+1}"
            d[key] = r[i]
        word = str(d.get("Ord") or d.get("ord") or "").strip()
        if not word:
            continue
        level = d.get("Niveau (1-30)") or d.get("Niveau") or d.get("niveau")
        try:
            level_int = int(level)
        except Exception:
            try:
                level_int = int(re.findall(r"\d+", str(level))[0])
            except Exception:
                level_int = None

        rows.append({
            "id": len(rows) + 1,
            "ord": word,
            "niveau": level_int,
            "fase": d.get("Fase"),
            "bogstaver": d.get("Bogstaver"),
            "stavelser": d.get("Stavelser"),
            "lydrethed": d.get("Lydrethed"),
            "stavemoenster": d.get("Stavemønster"),
            "morfologi": d.get("Morfologi"),
            "ordklasse": d.get("Ordklasse"),
            "interessekategori": d.get("Interessekategori"),
            "ordblind_risiko": d.get("Ordblind-risiko (0-3)") or d.get("Ordblind-risiko"),
            "ordblind_type": d.get("Ordblind-type"),
            "hyppighed": d.get("Hyppighed"),
            "kommentar": d.get("Kommentar"),
            "raw": d
        })

    payload = {
        "version": args.version,
        "generated": datetime.date.today().isoformat(),
        "sheet": sheet.title,
        "count": len(rows),
        "words": rows
    }
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {len(rows)} words to {out_path}")

if __name__ == "__main__":
    main()
